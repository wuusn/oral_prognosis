import torch
import glob
from random import random, uniform
from torchvision import datasets, models, transforms
import torchvision.transforms.functional as TF
from torch.utils.data import Dataset, DataLoader
from PIL import Image, ImageFilter
import numpy as np
from multiprocessing import Pool
import girder_client
import numpy as np
from skimage.transform import resize

class TMADataset(Dataset):
    # can change this to an abstract class
    def __init__(self, names, base_dir, phase, augmentations):
        self.names = names
        self.base_dir = base_dir
        self.phase = phase
        self.augmentations = augmentations
        self.files = self.get_files()
        
    def __len__(self):
        return len(self.files)

    def __getitem__(self, i):
        im_path = self.files[i]
        im = Image.open(im_path)
        ori_im = im.copy()
        if self.phase == 'train':
            for augmen in self.augmentations:
                faugmen = getattr(self, augmen+'_transform')
                im = faugmen(im)

        tsr_im = self.im2tensor(im)
        if phase == 'test':
            return {'image': tsr_im, 'ori_im': ori_im, 'new_im': im}
        label = im_path.split('/')[-3]
        label = int(label)
        return {'image': tsr_im, 'label': label, 'ori_im': ori_im, 'new_im': im} # in val, new==ori

    def get_files(self):
        files = []
        for name in self.names:
            tmpfiles = glob.glob(f'{base_dir}/*/{name}/*.png')
            files.extend(tmpfiles)
        return files

    def im2tensor(self, im):
        np_im = np.array(im)
        np_im = np_im.transpose((2,0,1))
        np_im = np_im / 255
        tsr_im = torch.from_numpy(np_im)
        return tsr_im

    def flip_transform(self, im):
        if random() > .5:
            im = TF.hflip(im)
        if random() > .5:
            im = TF.vflip(im)
        return im

    def simple_color_transform(self, im):
        if random() > .5:
            im = TF.adjust_brightness(im, uniform(1, 1.4))
        if random() > .5:
            im = TF.adjust_contrast(im, uniform(1, 1.4))
        if random() > .5:
            im = TF.adjust_saturation(im, uniform(1, 1.4))
        if random() > .5:
            im = TF.adjust_hue(im, uniform(-.5, .5))
        if random() > .5:
            im = im.filter(ImageFilter.GaussianBlur(int(random()>.5)+1))
        return im

if __name__ == '__main__':
    train_names = []
    from openpyxl import load_workbook
    fold_path = '/mnt/md0/_datasets/OralCavity/TMA_arranged/WU/data4disease/folds_info/round0.xlsx'
    wb = load_workbook(filename=fold_path)
    ws = wb['fold0']
    i=2
    while(ws[f'A{i}'].value!=None):
        train_names.append(ws[f'A{i}'].value)
        i+=1
    wb.close()

    base_dir = '/mnt/md0/_datasets/OralCavity/TMA_arranged/WU/data4disease/patch10x224s1.0e0.8'
    phase='train'
    augmens = ['flip', 'simple_color']

    dataset = TMADataset(train_names, base_dir, phase, augmens)
    dataset[1]



