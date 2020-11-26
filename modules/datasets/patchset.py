import torch
import glob
from random import random, uniform
from torchvision import datasets, transforms
import torchvision.transforms.functional as TF
from torch.utils.data import Dataset, DataLoader
from PIL import Image, ImageFilter
import numpy as np
from multiprocessing import Pool
import girder_client
import numpy as np
from skimage.transform import resize
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import math

class PatchDataset(Dataset):
    # can change this to an abstract class
    def __init__(self, names, base_dir, phase, augmentations):
        self.names = names
        self.base_dir = base_dir
        self.phase = phase
        self.augmentations = augmentations
        self.files = self.get_files()
        
    def __len__(self):
        return len(self.files)

    def __getitem__(self, i):
        im_path = self.files[i]
        im = Image.open(im_path)
        ori_im = im.copy()
        if self.phase == 'train':
            for augmen in self.augmentations:
                faugmen = getattr(self, augmen+'_transform')
                im = faugmen(im)

        tsr_im = self.im2tensor(im)
        ori_im=np.array(ori_im)
        im=np.array(im)
        if self.phase == 'test':
            return {'image': tsr_im, 'ori_im': ori_im, 'new_im': im}
        label = self.__get_label(im_path)
        return {'image': tsr_im, 'label': label, 'ori_im': ori_im, 'new_im': im} # in val, new==ori

    def __get_label(self, im_path):
        label = im_path.split('/')[-3]
        return int(label)

    def get_files(self):
        files = []
        for name in self.names:
            tmpfiles = glob.glob(f'{self.base_dir}/*/{name}/*.png')
            files.extend(tmpfiles)
        return files

    def im2tensor(self, im):
        np_im = np.array(im)
        np_im = np_im.transpose((2,0,1))
        np_im = np_im / 255
        tsr_im = torch.from_numpy(np_im)
        return tsr_im

    def flip_transform(self, im):
        if random() > .5:
            im = TF.hflip(im)
        if random() > .5:
            im = TF.vflip(im)
        return im

    def simple_color_transform(self, im):
        if random() > .5:
            im = TF.adjust_brightness(im, uniform(1, 1.4))
        if random() > .5:
            im = TF.adjust_contrast(im, uniform(1, 1.4))
        if random() > .5:
            im = TF.adjust_saturation(im, uniform(1, 1.4))
        if random() > .5:
            im = TF.adjust_hue(im, uniform(-.5, .5))
        if random() > .5:
            im = im.filter(ImageFilter.GaussianBlur(int(random()>.5)+1))
        return im

def get_names(fold_dir, r, f, phase):
    fold_path = f'{fold_dir}/round{r}.xlsx'
    wb = load_workbook(filename=fold_path)
    ws = wb[f'fold{f}']
    for i in range(1,ws.max_column+1):
        T = get_column_letter(i)
        if ws[f'{T}1'].value==phase:
            break
    i=2
    names=[]
    while(ws[f'{T}{i}'].value!=None):
        names.append(ws[f'{T}{i}'].value)
        i+=1
    wb.close()
    return names


def get_patchLoader(fold_dir, base_dir, r, f, phase, bs, augments):
    names = get_names(fold_dir,r,f,phase)
    augmens = augments.split(',')
    dataset = PatchDataset(names, base_dir, phase, augmens)
    shuffle = True if phase=='train' else False
    loader = DataLoader(dataset,batch_size=bs,shuffle=shuffle,num_workers=8,pin_memory=True)
    return loader

def get_slide_based_patchLoaders(fold_dir, base_dir, r,f,phase,bs,augments):
    names = get_names(fold_dir,r,f,phase)
    augmens = augments.split(',')
    loaders =[]
    for name in names:
        dataset = PatchDataset([name], base_dir, phase, augmens)
        shuffle = True if phase=='trian' else False
        loader = DataLoader(dataset,batch_size=bs,shuffle=shuffle,num_workers=8,pin_memory=True)
        loaders.append(loader)
    return loaders


if __name__ == '__main__':
    fold_xls_dir = '/mnt/md0/_datasets/OralCavity/TMA_arranged/WU/data4disease/folds_info'
    base_dir = '/mnt/md0/_datasets/OralCavity/TMA_arranged/WU/data4disease/patch10x224s1.0e0.8'
    r=0
    f=0
    augments='flip,simple_color'
    bs=32
    trainLoader = get_patchLoader(fold_xls_dir,base_dir,r,f,'train', bs,augments)
    for batch in trainLoader:
        size = batch['image'].size(0)
        if size != bs:
            print(size)
        continue



